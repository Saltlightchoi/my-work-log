import streamlit as st
import pandas as pd
from github import Github
import io
import re
import calendar
import os
from datetime import datetime, date
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import base64

# ==========================================
# 1. 환경 설정 및 전체 디자인 (CSS)
# ==========================================
st.set_page_config(layout="wide", page_title="장비 관리 통합 시스템")
st.markdown("""
    <style>
        .block-container { max-width: 98% !important; padding-top: 2rem !important; padding-bottom: 2rem !important; }
        [data-testid="stSidebar"] { width: 350px !important; }
        html, body, [class*="css"] { font-size: 16px !important; }
        .main-title { font-size: 2rem !important; font-weight: bold; padding-bottom: 15px !important; margin-top: -10px; }
        .info-box { background-color: #f8f9fa; padding: 15px; border-radius: 5px; border-left: 6px solid #4CAF50; margin-bottom: 15px; color: #333; font-size: 15px; }
        
        .final-report-table { width: 100%; border-collapse: collapse; border: 2px solid #000000 !important; font-size: 14px; color: #000000; background-color: #ffffff; }
        .final-report-table th, .final-report-table td { border: 1px solid #000000 !important; padding: 8px 10px; text-align: center; }
        .final-report-table th { background-color: #d9e1f2 !important; font-weight: bold; font-size: 15px; }
        .t-left { text-align: left !important; }
        div[data-testid="stSidebar"] button { width: 100% !important; font-weight: bold; font-size: 15px !important; }
    </style>
    """, unsafe_allow_html=True)

BASE_PATH_RAW = r"\\192.168.0.100\500 생산\550 국내CS\공유사진\\"
EQUIPMENT_OPTIONS = ["SLH1", "4010H", "3208H", "3208AT", "3208M", "3208C", "3208CM", "3208XM", "ADC200", "ADC300", "ADC400", "AH5200", "AM5"]

CS_TEMPLATE = [
    {"대항목": "공통", "순서": 1, "작업내용": "I/O Check\n- Out Put으로 동작 후 In Put LED 확인\n- Cylinder 정상 동작 확인\n- Manual에서 Cylinder 동작 후 LED 점등 확인\n- 미비된 부분 I/O List, PC에 저장 후 전장 수정 요청 진행\n- 전장 수정 후 수정되었는지 동작, LED 확인", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 2, "작업내용": "공압 Leak Check", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 3, "작업내용": "Cylinder Speed 조정 및 Part 위치 조정", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 4, "작업내용": "전면부 후면부 1차 Levelling\n- Auto Leveler 사용\n- Stacker Base 상단 -> 바닥면 400mm", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Stacker", "순서": 1, "작업내용": "L/D 1,2, UL/D 1,2, Emt, Reject Inverter 값 설정\n- 운전 모드 변경 P79-2(인터락 해제), P79-1(인터락 작동)\n- P3: 60, P4: 60(고속), P5: 30(중속), P6: 10(저속), P7: 5(가속), P8: 0(감속)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 1, "작업내용": "Motor Parameter Setting 및 다회전 클리어\n- S/W 팀 요청\n- 다회전 클리어 방법: Panaterm Ver.6.0 다운 후 해당 Parameter Servo Drive에 케이블 연결 -> 앰프 접속 -> 확인 -> 모니터 -> 다회전 클리어 클릭 -> 완료", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 2, "작업내용": "Precizer Up Rod 길이 변경 (57mm)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Hand", "순서": 1, "작업내용": "Loader Unloader X,Y축 직진도 (±0.5mm) Setting 및 측정\n- Dial Gauge Indicator 0.01mm(바늘)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Hand", "순서": 2, "작업내용": "L/D, UL/D Pitch, 높이(±0.15mm) Setting 및 측정\n- Dial Gauge Indicator 0.01mm(바늘)\n- Double Nut로 길이 조정", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Stacker", "순서": 1, "작업내용": "L/D 1,2, UL/D 1,2, Empty, Reject Base 평탄도(±0.2mm) Setting 및 측정\n- 디지털 전자 수평계 사용 X,Y축 ±0.2mm 이내", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Transfer", "순서": 1, "작업내용": "L/D, UL/D X,Y,Z축 직진도 평탄도(±0.3mm) Setting 및 측정\n- Dial Gauge Indicator 0.01mm(바늘)\n- 모든 무두 볼트 풀어놓은 후 측정 후 무드 볼트 조정\n- Z축 ±Limit Sensor 이동, Hard Stopper 위치변경", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Test", "순서": 1, "작업내용": "Front, Rear Press 평탄도(±0.25mm) Setting 및 측정\n- 디지털 거리 측정기, 측정 지그 사용\n- 무두 볼트 풀어놓은 후 측정 -> + 수치가 제일 큰 부분에 리셋 -> 무두 볼트 사용하여 평탄도 Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Test", "순서": 2, "작업내용": "Front, Rear Press Load Cell (0.2bar, 0.29bar) Setting 및 측정\n- Load Cell, 측정 지그\n- 언 컨텍 값(기구 이동 가능), 컨택값(해당 수치 도달 지점)\n- Load Cell 로드를 Match Plate X,Y 중간에서 컨택 해야함", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Test", "순서": 3, "작업내용": "Front, Rear T-Tray Rail 평탄도(±0.2mm) Setting 및 측정\n- 디지털 전자 수평계 사용\n- 측정 위치 : T-Tray 왼쪽 중간 오른쪽", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Set Plate", "순서": 1, "작업내용": "L/D 1,2, UL/D 1,2 X축 직진도(±0.2mm) (Gauge Block 사용) 평탄도(±0.2mm) Setting 및 측정\n- Dial Gauge Indicator 0.01mm(바늘)\n- Gauge Block 사용", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Stacker", "순서": 1, "작업내용": "Reject Front Rear Base 높이 조정\n- Rear Base Cylinder Rod 최대로 내린뒤, Front를 무두 볼트 사용하여 Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Set Plate", "순서": 1, "작업내용": "L/D 1,2, UL/D 1,2 Support Lock/Unlock Cylinder Rod Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Set Plate", "순서": 2, "작업내용": "L/D 1,2, UL/D 1,2 Down Hard Stopper 높이 Setting (24mm) 및 측정\n- Ex) L/D 1,2: 21mm, UL/D 1,2: 24mm Pass", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Middle", "순서": 1, "작업내용": "L/D, UL/D Gripper Up Cylinder Rod (37mm) Setting\n- Cylinder 상단에서 Joint Nut 상단 까지 37mm Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Input", "순서": 1, "작업내용": "Bottom Feeder Up Cylinder Rod (16.5mm) Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Output", "순서": 1, "작업내용": "Bottom Feeder Up Cylinder Rod (16.5mm) Setting", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 1, "작업내용": "전면부 후면부 Docking 및 Levelling\n- Auto Leveler 사용\n- Stacker Base 상단 -> 바닥면 400mm -> 전면부 풋 높이랑 동일하게 후면부 Setting\n- 전면부 Levelling -> Docking Pin 제거, 전면부 Cable Disconnect -> 후면부랑 전면부 Docking -> 후면부 Levelling -> Docking Pin 장착(안맞으면 빠루로 좌,우 이동) -> Cable Connect", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Middle", "순서": 1, "작업내용": "L/D, UL/D Rail Up/Down 높이, 평탄, 직진도 Setting (T-Tray 사용)\n- T-Tray 사용, Up/Down 상태에서 T-Tray가 흘려 내려가지 않게 Setting\n- Stopper 사용하여 Up/Down 높이 레벨 조정 (Up: 48.5mm, Down: 42.5mm)\n- Up 상태일때 후면부 Rail 한쪽이 안맞는다면 후면부 쪽 Rail 위치 조정 필요", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 1, "작업내용": "배출 Fan 장착 및 전원 Connect 연결\n- I/O 번호 확인, OS에서 Fan Check 확인", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Hand", "순서": 1, "작업내용": "L/D, UL/D X,Y축 반복도(±0.05mm) 측정 (10회)\n- Dial Gauge Indicator 0.01mm(바늘)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Transfer", "순서": 1, "작업내용": "L/D, UL/D X,Z축 반복도(±0.05mm) 측정 (10회)\n- Digital Indicator (0.001mm) 사용", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "Input", "순서": 1, "작업내용": "Bottom Feeder X축 반복도(±0.05mm) 측정 (10회)\n- Dial Gauge Indicator 0.01mm(바늘)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 1, "작업내용": "Transfer, Hand, Middle Feeder, Input, Output Motor Teaching", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 2, "작업내용": "후면부 덕트 -> 측면(Door) -> 후면(Door) -> 상부 Cover 장착 -> 부속 Cover 장착, Door Key, Door Sensor 장착 및 OS 확인", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 3, "작업내용": "Ionizer Loader A-01, Unloader A-02 설정", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 4, "작업내용": "접지 연결 및 확인", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 5, "작업내용": "환경 검수 List 확인 및 품질팀 Support", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 6, "작업내용": "Initial 및 Long Run 진행 (T-Tray 미사용)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 7, "작업내용": "Dry Run 진행 (T-Tray 사용)", "상태": "⬜ 대기", "비고": "", "첨부": ""},
    {"대항목": "공통", "순서": 8, "작업내용": "Dummy Run 진행 및 Clear Alarm (3000회)", "상태": "⬜ 대기", "비고": "", "첨부": ""}
]

# ==========================================
# 2. 데이터 관리 코어
# ==========================================
class DataManager:
    def __init__(self, repo, file_path, text_columns):
        self.repo = repo
        self.file_path = file_path
        self.text_columns = text_columns

    def load(self):
        try:
            file_content = self.repo.get_contents(self.file_path)
            df = pd.read_csv(io.StringIO(file_content.decoded_content.decode('utf-8-sig')))
            df = df.loc[:, ~df.columns.duplicated()]
            for col in self.text_columns:
                if col not in df.columns: df[col] = ""
                else: df[col] = df[col].fillna("").astype(str)
            return df, file_content.sha
        except:
            return pd.DataFrame(columns=self.text_columns), None

    def save(self, df, sha, message):
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        if sha: self.repo.update_file(self.file_path, message, csv_buffer.getvalue(), sha)
        else: self.repo.create_file(self.file_path, "Init Creation", csv_buffer.getvalue())

def maintain_project_order(df, original_order):
    df['__proj_cat__'] = pd.Categorical(df['프로젝트명'], categories=original_order, ordered=True)
    return df.sort_values(by=['__proj_cat__'], kind='stable').drop(columns=['__proj_cat__']).reset_index(drop=True)

def get_row_color(row):
    val = row.get('상태', '')
    if val == '✅ 완료': return ['background-color: rgba(76, 175, 80, 0.2)'] * len(row)
    elif val == '⏳ 작업중': return ['background-color: rgba(255, 193, 7, 0.2)'] * len(row)
    elif val == '🚨 보류': return ['background-color: rgba(244, 67, 54, 0.2)'] * len(row)
    return [''] * len(row)

# ==========================================
# 3. 화면 UI - 1페이지: 팀 업무일지
# ==========================================
def render_work_log_page(db_log):
    df_log, sha_log = db_log.load()
    if not df_log.empty and '날짜' in df_log.columns:
        df_log['날짜'] = pd.to_datetime(df_log['날짜']).dt.date.astype(str)
        df_log = df_log.sort_values(by='날짜', ascending=False).reset_index(drop=True)

    st.sidebar.markdown("---")
    st.sidebar.markdown("### 📝 일지 작성/수정/삭제")
    mode = st.sidebar.selectbox("기능 선택", ["➕ 작성", "✏️ 수정", "❌ 삭제"])
    
    if mode == "➕ 작성":
        with st.sidebar.form("add_log", clear_on_submit=True):
            d = st.date_input("날짜", datetime.today())
            e = st.selectbox("장비", EQUIPMENT_OPTIONS)
            c = st.text_area("업무 내용", height=300)
            n = st.text_input("비고")
            a = st.text_input("첨부 (FTP 경로 등)")
            if st.form_submit_button("저장하기"):
                new_row = pd.DataFrame([{"날짜": str(d), "장비": e, "작성자": st.session_state['user_name'], "업무내용": c, "비고": n, "첨부": a}])
                db_log.save(pd.concat([df_log, new_row], ignore_index=True), sha_log, f"Add Log: {d}")
                st.rerun()

    elif mode == "✏️ 수정" and not df_log.empty:
        idx = st.sidebar.selectbox("수정 대상", df_log.index, format_func=lambda x: f"{df_log.loc[x, '날짜']} | {df_log.loc[x, '업무내용'][:15]}...")
        with st.sidebar.form("edit_log"):
            e_date = st.date_input("날짜 수정", pd.to_datetime(df_log.loc[idx, '날짜']))
            e_content = st.text_area("내용 수정", value=df_log.loc[idx, '업무내용'], height=300)
            
            val_note = df_log.loc[idx, '비고'] if '비고' in df_log.columns else ""
            val_note = "" if pd.isna(val_note) else str(val_note)
            e_note = st.text_input("비고 수정", value=val_note)
            
            val_attach = df_log.loc[idx, '첨부'] if '첨부' in df_log.columns else ""
            val_attach = "" if pd.isna(val_attach) else str(val_attach)
            e_attach = st.text_input("첨부 수정", value=val_attach)
            
            if st.form_submit_button("수정 완료"):
                df_log.loc[idx, ['날짜', '업무내용', '비고', '첨부']] = [str(e_date), e_content, e_note, e_attach]
                db_log.save(df_log, sha_log, "Edit Log")
                st.rerun()

    elif mode == "❌ 삭제" and not df_log.empty:
        idx = st.sidebar.selectbox("삭제 대상 선택", df_log.index, format_func=lambda x: f"{df_log.loc[x, '날짜']} | {df_log.loc[x, '업무내용'][:10]}")
        st.sidebar.warning(f"내용: {df_log.loc[idx, '업무내용'][:50]}...")
        if st.sidebar.button("🗑️ 최종 삭제 (복구 불가)", type="primary"):
            db_log.save(df_log.drop(idx), sha_log, "Delete Log")
            st.rerun()

    col_title, col_excel = st.columns([8.5, 1.5])
    with col_title:
        st.markdown("<div class='main-title'>📝 팀 업무일지 대시보드</div>", unsafe_allow_html=True)
    with col_excel:
        csv_data = df_log.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
        st.download_button(label="📥 엑셀 다운로드", data=csv_data, file_name=f"work_log_{datetime.now().strftime('%Y%m%d')}.csv", use_container_width=True)

    st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
    
    filter_col1, filter_col2 = st.columns([2, 8])
    with filter_col1:
        equip_filter = st.selectbox("📌 장비모델 필터", ["전체"] + EQUIPMENT_OPTIONS)
    with filter_col2:
        search = st.text_input("🔍 내용/작성자 검색", placeholder="검색어를 입력하세요...")
        
    disp = df_log.copy()
    
    if equip_filter != "전체":
        disp = disp[disp['장비'].astype(str) == equip_filter]
        
    if search: 
        disp = disp[disp.apply(lambda r: search.lower() in str(r).lower(), axis=1)]
    
    st.dataframe(
        disp, 
        use_container_width=True, 
        hide_index=True, 
        column_config={
            "업무내용": st.column_config.TextColumn("업무내용", width="large"),
            "비고": st.column_config.TextColumn("비고", width="small"),
            "첨부": st.column_config.TextColumn("첨부", width="small")
        }
    )

# ==========================================
# 4. 화면 UI - 2페이지: CS 작업체크시트
# ==========================================
def render_cs_flow_page(db_flow):
    df_flow, sha_flow = db_flow.load()
    st.markdown("<div class='main-title'>✅ CS 작업 체크 시트</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)

    project_list = df_flow["프로젝트명"].unique().tolist() if not df_flow.empty else []
    
    if 'current_proj' not in st.session_state:
        st.session_state['current_proj'] = project_list[0] if project_list else ""
    if st.session_state['current_proj'] not in project_list:
        st.session_state['current_proj'] = project_list[0] if project_list else ""
    
    progress_dict = {}
    if not df_flow.empty:
        for proj in project_list:
            p_df = df_flow[df_flow["프로젝트명"] == proj]
            total_items = len(p_df)
            completed_items = len(p_df[p_df["상태"] == "✅ 완료"])
            pct = int((completed_items / total_items) * 100) if total_items > 0 else 0
            blocks = pct // 10
            bar = "🟩" * blocks + "⬜" * (10 - blocks)
            batt_icon = "🔋" if pct >= 20 else "🪫"
            progress_dict[proj] = f"{proj}  |  {batt_icon} {pct}% ({completed_items}/{total_items}) [{bar}]"

    col_a, col_b = st.columns(2)
    with col_a:
        with st.expander("➕ 새 프로젝트(호기) 시작하기"):
            with st.form("new_proj_form", clear_on_submit=True):
                new_proj = st.text_input("새 프로젝트명 (예: 4010H #2호기)")
                source_options = ["기본 템플릿(초기화 상태)"] + project_list
                source_proj = st.selectbox("어떤 형식을 복사할까요?", source_options, format_func=lambda x: progress_dict.get(x, x))
                if st.form_submit_button("프로젝트 생성하기") and new_proj and new_proj not in project_list:
                    if source_proj == "기본 템플릿(초기화 상태)": new_df = pd.DataFrame(CS_TEMPLATE)
                    else:
                        new_df = df_flow[df_flow["프로젝트명"] == source_proj].copy()
                        new_df[["상태", "비고", "첨부", "업데이트일"]] = ["⬜ 대기", "", "", ""]
                    new_df["프로젝트명"] = new_proj
                    db_flow.save(pd.concat([df_flow, new_df], ignore_index=True), sha_flow, f"Create: {new_proj}")
                    st.session_state['current_proj'] = new_proj
                    st.rerun()

    if project_list:
        sel_col, empty_col, save_col, del_col = st.columns([6, 2, 1, 1])
        default_idx = project_list.index(st.session_state['current_proj']) if project_list else 0
        with sel_col: 
            selected_proj = st.selectbox("📌 프로젝트 선택", project_list, index=default_idx, format_func=lambda x: progress_dict.get(x, x))
            st.session_state['current_proj'] = selected_proj 
        with save_col: 
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            btn_save = st.button("💾 변경 저장", use_container_width=True)
        with del_col: 
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if st.button("🗑️ 삭제", use_container_width=True):
                st.session_state['delete_target_proj'] = selected_proj
                st.rerun()

        mask = df_flow["프로젝트명"] == selected_proj
        proj_df = df_flow[mask].copy()

        if st.session_state.get('delete_target_proj') == selected_proj:
            st.error(f"🚨 [{selected_proj}] 프로젝트를 영구 삭제하시겠습니까?")
            if st.button("⚠️ 삭제 확정", type="primary"):
                db_flow.save(df_flow[~mask], sha_flow, f"Delete: {selected_proj}")
                st.session_state['delete_target_proj'] = None
                st.session_state['current_proj'] = project_list[0] if len(project_list) > 1 else ""
                st.rerun()
            if st.button("❌ 취소"): st.session_state['delete_target_proj'] = None; st.rerun()
            st.stop() 

        cats = proj_df['대항목'].unique().tolist()
        with st.expander("⚙️ 프로젝트 대항목 관리 (추가/수정/삭제/순서변경)"):
            c1, c2, c3, c4 = st.tabs(["➕ 대항목 추가", "✏️ 이름 수정", "❌ 대항목 삭제", "↕️ 순서 변경"])
            
            with c1:
                with st.form("add_cat_form", clear_on_submit=True):
                    new_c = st.text_input("새 대항목 이름")
                    if st.form_submit_button("추가하기"):
                        if new_c and new_c not in cats:
                            new_row = pd.DataFrame([{"프로젝트명": selected_proj, "대항목": new_c, "순서": 1, "작업내용": "새 작업 내용 입력", "상태": "⬜ 대기", "비고": "", "첨부": "", "업데이트일": ""}])
                            db_flow.save(pd.concat([df_flow, new_row], ignore_index=True), sha_flow, f"Add Cat: {new_c}")
                            st.success(f"'{new_c}' 항목이 추가되었습니다.")
                            st.rerun()

            with c2:
                with st.form("edit_cat_form", clear_on_submit=True):
                    target_c = st.selectbox("수정할 대항목 선택", cats)
                    rename_c = st.text_input("새로운 이름 입력")
                    if st.form_submit_button("이름 변경"):
                        if rename_c and rename_c not in cats:
                            df_flow.loc[(df_flow["프로젝트명"] == selected_proj) & (df_flow["대항목"] == target_c), "대항목"] = rename_c
                            db_flow.save(df_flow, sha_flow, f"Rename Cat: {target_c}")
                            st.success("이름이 변경되었습니다.")
                            st.rerun()

            with c3:
                with st.form("del_cat_form"):
                    del_c = st.selectbox("삭제할 대항목 선택", cats)
                    st.warning("⚠️ 해당 대항목과 안에 포함된 모든 세부 작업이 영구 삭제됩니다.")
                    if st.form_submit_button("삭제 실행"):
                        df_flow = df_flow[~((df_flow["프로젝트명"] == selected_proj) & (df_flow["대항목"] == del_c))]
                        db_flow.save(df_flow, sha_flow, f"Delete Cat: {del_c}")
                        st.success("삭제되었습니다.")
                        st.rerun()

            with c4:
                st.write("표 안의 **'새 순서'** 숫자를 클릭하여 순서를 변경하고 적용 버튼을 누르세요.")
                order_df = pd.DataFrame({"대항목": cats, "새 순서": range(1, len(cats)+1)})
                edited_order = st.data_editor(order_df, hide_index=True, use_container_width=True)
                if st.button("변경된 순서 적용하기"):
                    edited_order = edited_order.sort_values("새 순서")
                    ordered_cats = edited_order["대항목"].tolist()
                    proj_df['__cat_order__'] = pd.Categorical(proj_df['대항목'], categories=ordered_cats, ordered=True)
                    sorted_proj_df = proj_df.sort_values(['__cat_order__', '순서']).drop(columns=['__cat_order__'])
                    
                    new_df_flow = df_flow[df_flow["프로젝트명"] != selected_proj]
                    new_df_flow = pd.concat([new_df_flow, sorted_proj_df], ignore_index=True)
                    db_flow.save(new_df_flow, sha_flow, "Reorder Cats")
                    st.success("순서가 적용되었습니다.")
                    st.rerun()

        total_tasks = len(proj_df); comp_tasks = len(proj_df[proj_df["상태"] == "✅ 완료"])
        pct_float = (comp_tasks / total_tasks) if total_tasks > 0 else 0.0
        st.markdown(f"<div style='font-size:16px; font-weight:bold; color:#4CAF50;'>⚡ 전체 진행도 ({comp_tasks} / {total_tasks})</div>", unsafe_allow_html=True)
        st.progress(pct_float, text=f"{int(pct_float * 100)}% 완료")

        proj_df['group_id'] = (proj_df['대항목'] != proj_df['대항목'].shift()).cumsum()
        groups = proj_df.groupby('group_id', sort=False) 
        edited_dfs = []; current_user_stamp = f"{st.session_state['user_name']} ({datetime.today().strftime('%y-%m-%d')})"

        for group_id, group_df in groups:
            cat = group_df['대항목'].iloc[0]
            display_df = group_df.drop(columns=['group_id']).reset_index(drop=True)
            curr_stats = display_df['상태'].tolist()
            
            cat_total = len(curr_stats)
            cat_comp = curr_stats.count('✅ 완료')
            cnt_str = f"({cat_comp}/{cat_total})"
            
            if '🚨 보류' in curr_stats: tab_title = f"🔴 [보류] {cat} {cnt_str}"
            elif curr_stats and all(s == '✅ 완료' for s in curr_stats): tab_title = f"🟢 [완료] {cat} {cnt_str}"
            elif any(s in ['⏳ 작업중', '✅ 완료'] for s in curr_stats): tab_title = f"🟡 [진행] {cat} {cnt_str}"
            else: tab_title = f"📍 [대기] {cat} {cnt_str}"
            
            with st.expander(tab_title, expanded=False):
                styled_df = display_df.style.apply(get_row_color, axis=1)
                edited_cat_df = st.data_editor(
                    styled_df, use_container_width=True, hide_index=True, num_rows="dynamic", key=f"editor_{selected_proj}_{group_id}",
                    column_config={
                        "순서": st.column_config.NumberColumn("No", width="small"), 
                        "작업내용": st.column_config.TextColumn("세부 작업 내용", width="large"), 
                        "상태": st.column_config.SelectboxColumn("상태", options=["⬜ 대기", "⏳ 작업중", "✅ 완료", "🚨 보류"], width="small"),
                        "비고": st.column_config.TextColumn("비고", width="small"),
                        "첨부": st.column_config.TextColumn("첨부", width="small")
                    }
                )
                for idx, new_row in edited_cat_df.iterrows():
                    if new_row['상태'] == "⬜ 대기": edited_cat_df.at[idx, '업데이트일'] = ""
                    else:
                        match = display_df[(display_df['작업내용'] == new_row['작업내용']) & (display_df['상태'] == new_row['상태']) & (display_df['비고'] == new_row['비고'])]
                        if match.empty: edited_cat_df.at[idx, '업데이트일'] = current_user_stamp
                        else: edited_cat_df.at[idx, '업데이트일'] = match.iloc[0]['업데이트일']
                edited_cat_df["대항목"] = cat; edited_cat_df["프로젝트명"] = selected_proj; edited_cat_df["org_group_id"] = group_id 
                edited_dfs.append(edited_cat_df)

        if btn_save:
            updated_proj_df = pd.concat(edited_dfs, ignore_index=True)
            if not updated_proj_df.empty:
                updated_proj_df = updated_proj_df.sort_values(by=['org_group_id', '순서'], kind='stable')
                updated_proj_df['group_id'] = (updated_proj_df['대항목'] != updated_proj_df['대항목'].shift()).cumsum()
                updated_proj_df["순서"] = updated_proj_df.groupby('group_id').cumcount() + 1
                updated_proj_df = updated_proj_df.drop(columns=['group_id', 'org_group_id']).reset_index(drop=True)
            original_projects = df_flow['프로젝트명'].unique().tolist()
            new_df_flow = pd.concat([df_flow[~mask], updated_proj_df], ignore_index=True)
            db_flow.save(maintain_project_order(new_df_flow, original_projects), sha_flow, f"Update: {selected_proj}")
            st.success("✅ 저장되었습니다!"); st.rerun()
    else: st.info("프로젝트가 없습니다.")

# ==========================================
# 5. 화면 UI - 3페이지: 장비 가동 데이터 
# ==========================================
def render_equipment_data_page():
    import re
    from plotly.subplots import make_subplots
    import pandas as pd

    st.markdown("""
        <style>
            .final-report-table {
                width: 100%; border-collapse: collapse; border: 2px solid #000000 !important;
                font-size: 12px; color: #000000; background-color: #ffffff;
            }
            .final-report-table th, .final-report-table td {
                border: 1px solid #000000 !important; padding: 6px 8px; text-align: center !important;
            }
            .final-report-table th { background-color: #d9e1f2 !important; font-weight: bold; }
            .t-left { text-align: left !important; }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("<div class='main-title'>📊 장비 가동 데이터 정밀 분석</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1: equipment = st.selectbox("장비 선택", EQUIPMENT_OPTIONS, key="eq_data_equip")
    with col2: unit = st.selectbox("호기 선택", ["1호기", "2호기", "3호기", "4호기", "5호기"], key="eq_data_unit")
    with col3: month_str = st.selectbox("조회할 월 선택", ["1월", "2월", "3월"], key="eq_data_month")

    file_map = {"1월": "SLH1 - January 2026.xlsx", "2월": "SLH1 - February 2026.xlsx", "3월": "SLH1 - March 2026.xlsx"}
    target_file = file_map.get(month_str, "")
    month_num = month_str.replace("월", "")

    try:
        xls = pd.read_excel(target_file, sheet_name=None, header=None, engine='openpyxl')
        df_raw = None
        for name, data in xls.items():
            if data.astype(str).apply(lambda r: r.str.contains('Unit|Output', case=False).any(), axis=1).any():
                df_raw = data; break
        
        if df_raw is None:
            st.error("⚠️ 가동 데이터를 찾을 수 없습니다."); return

        def get_sum_row(keywords):
            for _, row in df_raw.iterrows():
                row_str = "".join(row.astype(str)).lower().replace(" ", "").replace("#", "").replace("_", "")
                if any(k in row_str for k in keywords) and not any(x in row_str for x in ['%', '발생률']):
                    vals = row.tolist()
                    for i, v in enumerate(vals):
                        v_s = str(v).replace('.','').replace(',','').strip()
                        if v_s.isdigit() or v_s in ['비가동', '미가동']: return (vals[i:i+31]+[0]*31)[:31]
            return [0]*31

        units = get_sum_row(['totalunit', 'output'])
        jams = get_sum_row(['jamcount', 'jam'])
        ppjs = get_sum_row(['ppj'])

        chart_df = pd.DataFrame({'날짜': [f"{month_num}/{i}" for i in range(1, 32)], 'Unit': units, 'Jam': jams, 'PPJ': ppjs})
        for c in ['Unit', 'Jam', 'PPJ']:
            chart_df[c] = pd.to_numeric(chart_df[c].astype(str).str.replace(',', '').replace(['nan','비가동','미가동','None',''], '0'), errors='coerce').fillna(0)

        chart_df['Cum_Unit'] = chart_df['Unit'].cumsum()
        chart_df['Cum_Jam'] = chart_df['Jam'].cumsum()
        chart_df['Cum_PPJ'] = chart_df.apply(lambda row: round(row['Cum_Unit'] / row['Cum_Jam'], 1) if row['Cum_Jam'] > 0 else 0, axis=1)

        fig = make_subplots(rows=2, cols=1, shared_xaxes=True, vertical_spacing=0.12,
                            subplot_titles=("투입량(Unit) 및 에러(Jam) 건수", "생산 효율(PPJ) 추이"),
                            specs=[[{"secondary_y": True}], [{"secondary_y": False}]])
        
        fig.add_trace(go.Bar(x=chart_df['날짜'], y=chart_df['Unit'], name='투입(Unit)', marker_color='#5B9BD5'), row=1, col=1, secondary_y=False)
        fig.add_trace(go.Scatter(x=chart_df['날짜'], y=chart_df['Jam'], name='에러(Jam)', mode='lines+markers', line=dict(color='#ED7D31', width=2)), row=1, col=1, secondary_y=True)
        
        fig.add_trace(go.Bar(x=chart_df['날짜'], y=chart_df['PPJ'], name='일별 PPJ', marker_color='#A9D18E', opacity=0.8), row=2, col=1)
        fig.add_trace(go.Scatter(x=chart_df['날짜'], y=chart_df['Cum_PPJ'], name='월 누적 PPJ (실선)', mode='lines+markers', line=dict(color='#FF0000', width=4)), row=2, col=1)
        
        fig.update_layout(height=650, margin=dict(l=50, r=50, t=50, b=50), hovermode="x unified", showlegend=True)
        fig.update_yaxes(title_text="투입량 (EA)", secondary_y=False, row=1, col=1)
        fig.update_yaxes(title_text="Jam (건)", secondary_y=True, row=1, col=1)
        fig.update_yaxes(title_text="PPJ", row=2, col=1)
        st.plotly_chart(fig, use_container_width=True)

        st.subheader(f"📋 {month_str} 에러 상세 리스트")
        h_idx = -1
        for i, row in df_raw.iterrows():
            row_str = "".join(row.astype(str)).lower()
            if 'error code' in row_str or 'errorcode' in row_str:
                h_idx = i; h_row = row.tolist(); break

        if h_idx != -1:
            m = {'D': None, 'C': None, 'M': None, 'A': None, 'T': None, 'L': None, 'P': None}
            for i, v in enumerate(h_row):
                v_str = str(v).lower().replace(' ', '').replace('\n', '')
                if m['D'] is None and 'date' in v_str: m['D'] = i
                elif m['C'] is None and 'errorcode' in v_str: m['C'] = i
                elif m['M'] is None and ('massage' in v_str or 'message' in v_str): m['M'] = i
                elif m['A'] is None and ('finding' in v_str or 'action' in v_str): m['A'] = i
                elif m['T'] is None and 'time' in v_str: m['T'] = i
                elif m['L'] is None and 'point' in v_str: m['L'] = i
                elif m['P'] is None and 'ppj' in v_str: m['P'] = i

            data_slice = df_raw.iloc[h_idx + 1:].copy()
            cleaned_list = []
            
            def is_meaningful(val):
                if pd.isna(val): return False
                cleaned = re.sub(r'[^a-zA-Z0-9가-힣]', '', str(val))
                return len(cleaned) > 0

            for _, r in data_slice.iterrows():
                has_code = is_meaningful(r[m['C']]) if m['C'] is not None else False
                has_msg = is_meaningful(r[m['M']]) if m['M'] is not None else False
                
                if not has_code and not has_msg:
                    continue
                
                code_orig = str(r[m['C']]).strip() if m['C'] is not None else ""
                if code_orig.endswith('.0'): code_orig = code_orig[:-2]
                
                msg_orig = str(r[m['M']]).strip() if m['M'] is not None else ""
                if msg_orig.lower() in ['nan', 'none']: msg_orig = ""

                dt = None
                if m['D'] is not None:
                    raw_d = r[m['D']]
                    if is_meaningful(raw_d):
                        if str(raw_d).replace('.','').isdigit():
                            dt = pd.to_datetime(float(raw_d), unit='D', origin='1899-12-30').strftime('%Y-%m-%d')
                        else:
                            dt = str(raw_d).split(' ')[0]

                ppj_val = None
                if m['P'] is not None and m['P'] < len(r):
                    raw_p = str(r[m['P']]).strip()
                    if is_meaningful(raw_p) and raw_p.lower() not in ['nan', 'none']:
                        ppj_val = raw_p.split('.')[0]

                t_val = ""
                if m['T'] is not None and m['T'] < len(r):
                    raw_t = str(r[m['T']]).strip()
                    if is_meaningful(raw_t) and raw_t.lower() not in ['nan', 'none']: 
                        t_val = raw_t.split('.')[0] if ':' in raw_t else raw_t
                
                a_val = ""
                if m['A'] is not None and m['A'] < len(r):
                    raw_a = str(r[m['A']]).strip()
                    if is_meaningful(raw_a) and raw_a.lower() not in ['nan', 'none']: a_val = raw_a

                l_val = ""
                if m['L'] is not None and m['L'] < len(r):
                    raw_l = str(r[m['L']]).strip()
                    if is_meaningful(raw_l) and raw_l.lower() not in ['nan', 'none']: l_val = raw_l

                cleaned_list.append({
                    "Date": dt, "Code": code_orig, "PPJ": ppj_val,
                    "Msg": msg_orig, "Act": a_val, "Time": t_val, "Loc": l_val
                })

            if cleaned_list:
                final_df = pd.DataFrame(cleaned_list)
                final_df['Date'] = final_df['Date'].replace(['', 'nan', 'None'], pd.NA).ffill()
                final_df['PPJ'] = final_df['PPJ'].replace(['', 'nan', 'None'], pd.NA).ffill().fillna("0")
                
                rows_html = ""
                for _, row in final_df.iterrows():
                    date_td = f"{row['Date']}" if not pd.isna(row['Date']) else ""
                    rows_html += f"<tr><td style='width:75px;'>{date_td}</td><td style='width:60px;'>{row['Code']}</td><td style='width:60px;'>{row['PPJ']}</td><td class='t-left'>{row['Msg']}</td><td class='t-left'>{row['Act']}</td><td style='width:65px;'>{row['Time']}</td><td style='width:90px;'>{row['Loc']}</td></tr>"
                
                st.markdown(f"<table class='final-report-table'><thead><tr><th style='width:75px;'>날짜</th><th style='width:60px;'>코드</th><th style='width:60px;'>PPJ</th><th>에러내용</th><th>조치내용</th><th style='width:65px;'>시간</th><th style='width:90px;'>위치</th></tr></thead><tbody>{rows_html}</tbody></table>", unsafe_allow_html=True)
            else: st.info("상세 데이터가 없습니다.")
        else: st.info("데이터 헤더를 찾을 수 없습니다.")

    except Exception as e: st.error(f"⚠️ 시스템 오류: {e}")

# ==========================================
# 6. 화면 UI - 4페이지: ECN & STN
# ==========================================
def render_ecn_stn_page(repo):
    st.markdown("<div class='main-title'>🛠️ ECN & STN (장비 파트 및 수정사항 관리)</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)

    # 오타 수정 완료: ECT -> ECN
    ecn_base_path = r"\\192.168.0.100\500 생산\550 국내CS\ECN&STN"

    col1, col2, col3, col_search = st.columns([1.5, 1.5, 2.5, 4.5])
    with col1: equipment = st.selectbox("장비 선택", EQUIPMENT_OPTIONS, key="ecn_equip")
    with col2: unit = st.selectbox("호기 선택", ["전체"] + [f"{i}호기" for i in range(1, 16)], key="ecn_unit")
    
    with col3:
        st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
        show_help = st.checkbox("💡 도움말 및 수정방법 보기")
            
    with col_search:
        st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
        search_keyword = st.text_input("🔍 내용/ECN No. 검색", placeholder="예: ECN-005, 실린더 교체 등", label_visibility="collapsed")

    if show_help:
        st.info(f"**이용 안내:** 깃허브 `data/ECN/` 폴더 안의 **`ECN_STN_Master({equipment}).xlsx`** 파일을 기반으로 목록을 출력합니다.\n\n"
                f"표의 **'조치현황'**, **'특이사항'**, **'첨부(파일명 입력)'** 칸을 더블 클릭하여 내용을 직접 수정할 수 있습니다. 수정한 뒤엔 하단의 **저장 버튼**을 눌러주세요.\n\n"
                "**📁 첨부파일/원본 열기 팁:**\n"
                "웹 브라우저 표에서 복사할 때 따옴표가 3개(`\"\"\"`)로 증식하는 버그를 완벽히 피하기 위해 화면 하단에 **[1초 복사기]**를 만들었습니다.\n"
                f"1. 표의 **'첨부(파일명 입력)'** 칸에는 **파일 이름만** 입력하세요.\n"
                "2. 열고 싶은 항목의 파일 이름을 복사(`Ctrl+C`)합니다.\n"
                "3. 표 밑에 있는 **'1초 복사기'** 칸에 붙여넣기 하시면 완벽한 주소가 생성됩니다!\n"
                "4. 키보드에서 **`[윈도우키 + R]`**을 눌러 붙여넣고 엔터를 치면 파일이 바로 열립니다.")

    target_file = f"data/ECN/ECN_STN_Master({equipment}).xlsx"

    try:
        file_content = repo.get_contents(target_file)
        raw_bytes = file_content.decoded_content
        excel_data = io.BytesIO(raw_bytes)
        
        xls_dict = pd.read_excel(excel_data, engine='openpyxl', header=None, sheet_name=None)
        sheet_names = list(xls_dict.keys())
        first_sheet = sheet_names[0]
        df_raw = xls_dict[first_sheet]
        
        if df_raw.empty:
            st.warning("⚠️ 선택하신 엑셀 파일이 비어있습니다. 양식을 먼저 채워주세요.")
            return

        h_idx = -1
        for i, row in df_raw.head(20).iterrows():
            row_str = "".join(row.astype(str)).replace(" ", "").lower()
            if '날짜' in row_str and ('발행' in row_str or '장비호기' in row_str or '내용' in row_str or 'as-is' in row_str):
                h_idx = i
                break
        
        if h_idx != -1:
            df = df_raw.iloc[h_idx + 1:].reset_index(drop=True)
            orig_cols = df_raw.iloc[h_idx].tolist()
            df['Original_Index'] = range(h_idx + 1, len(df_raw))
        else:
            df = df_raw.iloc[1:].reset_index(drop=True)
            orig_cols = df_raw.iloc[0].tolist()
            df['Original_Index'] = range(1, len(df_raw))
            
        seen_cols = set()
        new_cols = []
        col_idx_map = {} 
        for i, c in enumerate(orig_cols):
            c_clean = str(c).replace(" ", "").upper()
            base_col = ""
            
            if '날짜' in c_clean or '일자' in c_clean: base_col = '날짜'
            elif '발행부서' in c_clean: base_col = '발행부서'
            elif '발행자' in c_clean or '작성자' in c_clean: base_col = '발행자'
            elif '장비호기' in c_clean or '호기' in c_clean: base_col = '장비호기'
            elif 'ECN' in c_clean or '문서번호' in c_clean: base_col = 'ECN No'
            elif 'AS-IS' in c_clean or 'ASIS' in c_clean or '내용' in c_clean: base_col = 'AS-IS'
            elif 'TO-BE' in c_clean or 'TOBE' in c_clean or '변경' in c_clean: base_col = 'TO-BE'
            elif '특이사항' in c_clean or '비고' in c_clean: 
                base_col = '특이사항'
                col_idx_map['특이사항'] = i
            elif '조치' in c_clean or '진행' in c_clean: 
                base_col = '조치현황'
                col_idx_map['조치현황'] = i
            elif '첨부' in c_clean: 
                base_col = '첨부'
                col_idx_map['첨부'] = i
            else: base_col = str(c).strip()

            if base_col in seen_cols:
                base_col = f"{base_col}_{i}"
            seen_cols.add(base_col)
            new_cols.append(base_col)
        
        new_cols.append('Original_Index')
        df.columns = new_cols
        
        if '장비호기' in df.columns:
            if unit == "전체":
                filtered_df = df.copy()
            else:
                target_match = re.search(r'(\d+)호기', unit)
                target_num = int(target_match.group(1)) if target_match else -1
                
                def check_match(val):
                    if pd.isna(val): return False
                    val_str = str(val).lower()
                    if unit.lower() in val_str: return True
                    ranges = re.findall(r'(\d+)\s*[~-]\s*(\d+)', val_str)
                    for s_str, e_str in ranges:
                        s, e = int(s_str), int(e_str)
                        if s > e: s, e = e, s
                        if s <= target_num <= e: return True
                    return False

                mask = df['장비호기'].apply(check_match)
                filtered_df = df[mask].copy()
        elif '발행부서' in df.columns: 
            filtered_df = df.copy()
        else:
            st.error("⚠️ 엑셀 파일 컬럼을 인식할 수 없습니다. 양식을 다시 확인해주세요.")
            return

        if search_keyword:
            filtered_df = filtered_df[filtered_df.apply(lambda r: search_keyword.lower() in str(r).lower(), axis=1)]
            
        expected_cols = ['Original_Index', '날짜', '발행부서', '발행자', 'ECN No', 'AS-IS', 'TO-BE', '특이사항', '조치현황', '첨부']
        display_cols = [c for c in expected_cols if c in filtered_df.columns]
        filtered_df = filtered_df[display_cols].copy()
        
        if '날짜' in filtered_df.columns:
            import datetime as dt
            def parse_date_robust(d):
                if isinstance(d, dt.time): return pd.NaT
                if isinstance(d, (dt.datetime, dt.date)): return pd.to_datetime(d)
                try:
                    if pd.isna(d): return pd.NaT
                except: pass
                d_str = str(d).strip()
                if d_str in ['', 'nan', 'NaN', 'None', 'nat', 'NaT', '0.0']: return pd.NaT
                if d_str.replace('.', '', 1).isdigit():
                    try:
                        val = float(d_str)
                        if 30000 < val < 80000: return pd.to_datetime(val, unit='D', origin='1899-12-30')
                    except: pass
                try: 
                    d_str_clean = d_str.replace('.', '-').replace('/', '-')
                    return pd.to_datetime(d_str_clean, errors='coerce')
                except: return pd.NaT
                    
            filtered_df['TempDate'] = filtered_df['날짜'].apply(parse_date_robust)
            filtered_df = filtered_df.dropna(subset=['TempDate'])
            filtered_df = filtered_df.sort_values(by='TempDate', ascending=False)
            filtered_df['날짜'] = filtered_df['TempDate'].dt.strftime('%Y-%m-%d')
            filtered_df = filtered_df.drop(columns=['TempDate'])
            
        filtered_df = filtered_df.astype(str).replace(['nan', 'NaN', 'None', 'nat', 'NaT', '0.0'], '')
        filtered_df.reset_index(drop=True, inplace=True)
        
        if '첨부' in filtered_df.columns:
            def clean_attachment(val):
                if pd.isna(val): return ""
                val_str = str(val).strip().replace('"', '') 
                if val_str.startswith(ecn_base_path):
                    return val_str[len(ecn_base_path):].lstrip("\\")
                return val_str
            
            filtered_df['첨부(파일명)'] = filtered_df['첨부'].apply(clean_attachment)
        
        st.markdown("<hr style='margin-top: 5px; margin-bottom: 15px;'>", unsafe_allow_html=True)
        if not filtered_df.empty:
            total_cnt = len(filtered_df)
            done_cnt = 0
            prog_cnt = 0
            if '조치현황' in filtered_df.columns:
                done_cnt = len(filtered_df[filtered_df['조치현황'].astype(str).str.contains('완료', na=False)])
                prog_cnt = len(filtered_df[filtered_df['조치현황'].astype(str).str.contains('진행중|진행 중', na=False)])
            pend_cnt = total_cnt - done_cnt - prog_cnt
            
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📌 검색된 총 건수", f"{total_cnt} 건")
            m2.metric("✅ 조치 완료", f"{done_cnt} 건")
            m3.metric("⏳ 진행중", f"{prog_cnt} 건")
            m4.metric("🚨 미조치 (대기)", f"{pend_cnt} 건")
        
        st.markdown("<br>", unsafe_allow_html=True)

        if not filtered_df.empty:
            disabled_cols = [c for c in filtered_df.columns if c not in ['특이사항', '조치현황', '첨부(파일명)']]
            
            def highlight_status(val):
                val_str = str(val).strip()
                if '완료' in val_str: return 'background-color: #d4edda; color: #155724; font-weight: bold;'
                if '진행' in val_str: return 'background-color: #fff3cd; color: #856404; font-weight: bold;'
                if '대기' in val_str or val_str == '': return 'background-color: #f8d7da; color: #721c24; font-weight: bold;'
                return ''

            if '조치현황' in filtered_df.columns:
                styled_df = filtered_df.style.map(highlight_status, subset=['조치현황'])
            else:
                styled_df = filtered_df

            col_cfg = {"Original_Index": None}
            if "첨부" in filtered_df.columns: col_cfg["첨부"] = None
            if "AS-IS" in filtered_df.columns: col_cfg["AS-IS"] = st.column_config.TextColumn("AS-IS", width="large")
            if "TO-BE" in filtered_df.columns: col_cfg["TO-BE"] = st.column_config.TextColumn("TO-BE", width="large")
            if "특이사항" in filtered_df.columns: col_cfg["특이사항"] = st.column_config.TextColumn("특이사항", width="medium")
            if "조치현황" in filtered_df.columns: col_cfg["조치현황"] = st.column_config.TextColumn("조치현황", width="small")
            
            # 안전장치: 표(엑셀)에 진짜로 이 열이 있을 때만 설정 적용!
            if "첨부(파일명)" in filtered_df.columns: 
                col_cfg["첨부(파일명)"] = st.column_config.TextColumn("첨부(파일명 입력)", width="medium")

            edited_df = st.data_editor(
                styled_df, 
                use_container_width=True, 
                hide_index=True,
                disabled=disabled_cols,
                column_config=col_cfg,
                key=f"ecn_editor_safe_{equipment}_{unit}_{search_keyword}"
            )
            
            # ★ 하단 복사기 부활: 웹 브라우저 따옴표 증식 버그 회피용!
            st.markdown("---")
            st.markdown("#### 🚀 원본 파일 바로 열기 (1초 복사기)")
            st.info("웹 표에서 복사하면 따옴표가 늘어나는 버그가 있어 만든 전용 복사기입니다. 위 표에서 **파일명만 복사**해서 아래에 붙여넣어 주세요.")
            
            run_target = st.text_input("여기에 파일명을 붙여넣으세요:", placeholder="예: SLH1-PP-260306-01(5건).pdf", label_visibility="collapsed")
            if run_target:
                clean_target = run_target.strip().replace('"', '')
                if clean_target.startswith(ecn_base_path):
                    clean_target = clean_target[len(ecn_base_path):].lstrip("\\")
                final_run_path = f'"{ecn_base_path}\\{clean_target}"'
                
                st.success("✨ 변환 완료! 아래 회색 박스 우측 상단의 **[복사 아이콘(📋)]**을 클릭하고 `[Win + R]` 창에 붙여넣기 하세요.")
                st.code(final_run_path, language="text")
            
            action_col1, action_col2, action_col3 = st.columns([2, 2, 6])
            with action_col1:
                save_btn = st.button("💾 변경사항 엑셀에 자동 저장하기", type="primary", use_container_width=True)
            with action_col2:
                output_excel = io.BytesIO()
                with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
                    cols_to_drop = ['Original_Index', '첨부(파일명)']
                    filtered_df.drop(columns=cols_to_drop, errors='ignore').to_excel(writer, index=False, sheet_name='ECN_Data')
                st.download_button(
                    label="📥 현재 리스트 엑셀 다운로드",
                    data=output_excel.getvalue(),
                    file_name=f"ECN_{equipment}_{unit}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            with st.expander("➕ 새 ECN 항목 엑셀에 바로 등록하기 (웹 기반)"):
                with st.form("add_new_ecn_form", clear_on_submit=True):
                    st.write("아래 내용을 작성하여 등록하면 엑셀 파일 맨 아래에 자동으로 추가됩니다.")
                    f_col1, f_col2, f_col3, f_col4 = st.columns(4)
                    n_date = f_col1.date_input("날짜")
                    n_dept = f_col2.text_input("발행부서")
                    n_author = f_col3.text_input("발행자")
                    n_unit = f_col4.text_input("장비호기", value=f"{equipment} {unit if unit != '전체' else ''}")
                    
                    n_ecn = st.text_input("ECN No (예: ECN-001)")
                    f_col5, f_col6 = st.columns(2)
                    n_asis = f_col5.text_area("AS-IS (수정 전)")
                    n_tobe = f_col6.text_area("TO-BE (수정 후)")
                    
                    f_col7, f_col8, f_col9 = st.columns([2, 1, 2])
                    n_note = f_col7.text_input("특이사항")
                    n_status = f_col8.selectbox("조치현황", ["대기", "진행중", "완료"])
                    n_attach = f_col9.text_input("첨부 (파일명만 입력)", placeholder="예: SLH1-PP-260306-01.pdf")
                    
                    if st.form_submit_button("새 항목 등록하기"):
                        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
                        ws = wb[first_sheet]
                        new_row_data = [""] * len(orig_cols)
                        for i, c in enumerate(orig_cols):
                            c_clean = str(c).replace(" ", "").upper()
                            if '날짜' in c_clean or '일자' in c_clean: new_row_data[i] = str(n_date)
                            elif '발행부서' in c_clean: new_row_data[i] = n_dept
                            elif '발행자' in c_clean or '작성자' in c_clean: new_row_data[i] = n_author
                            elif '장비호기' in c_clean or '호기' in c_clean: new_row_data[i] = n_unit
                            elif 'ECN' in c_clean or '문서번호' in c_clean: new_row_data[i] = n_ecn
                            elif 'AS-IS' in c_clean or 'ASIS' in c_clean or '내용' in c_clean: new_row_data[i] = n_asis
                            elif 'TO-BE' in c_clean or 'TOBE' in c_clean or '변경' in c_clean: new_row_data[i] = n_tobe
                            elif '특이사항' in c_clean or '비고' in c_clean: new_row_data[i] = n_note
                            elif '조치' in c_clean or '진행' in c_clean: new_row_data[i] = n_status
                            elif '첨부' in c_clean: 
                                n_attach_clean = n_attach.strip().replace('"', '')
                                if n_attach_clean and not n_attach_clean.startswith("\\\\") and not n_attach_clean.startswith("http"):
                                    new_row_data[i] = f"{ecn_base_path}\\{n_attach_clean}"
                                else:
                                    new_row_data[i] = n_attach_clean
                        ws.append(new_row_data)
                        
                        output = io.BytesIO()
                        wb.save(output)
                        repo.update_file(target_file, f"Add new ECN via Web", output.getvalue(), file_content.sha)
                        st.success("✅ 새 ECN 항목이 추가되었습니다! 화면을 새로고침 합니다.")
                        st.rerun()

            if save_btn:
                try:
                    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
                    ws = wb[first_sheet]
                    changes_made = False
                    for _, row in edited_df.iterrows():
                        orig_idx = int(row['Original_Index'])
                        xl_row = orig_idx + 1 
                        if '특이사항' in col_idx_map:
                            ws.cell(row=xl_row, column=col_idx_map['특이사항'] + 1, value=row.get('특이사항', ''))
                            changes_made = True
                        if '조치현황' in col_idx_map:
                            ws.cell(row=xl_row, column=col_idx_map['조치현황'] + 1, value=row.get('조치현황', ''))
                            changes_made = True
                        if '첨부' in col_idx_map:
                            fname = str(row.get('첨부(파일명)', '')).strip().replace('"', '')
                            if fname and not fname.startswith("\\\\") and not fname.startswith("http"):
                                full_path = f"{ecn_base_path}\\{fname}"
                            else:
                                full_path = fname
                            ws.cell(row=xl_row, column=col_idx_map['첨부'] + 1, value=full_path)
                            changes_made = True
                            
                    if changes_made:
                        output = io.BytesIO()
                        wb.save(output)
                        repo.update_file(target_file, f"Dashboard Update: ECN ({equipment}-{unit})", output.getvalue(), file_content.sha)
                        st.success("✅ 엑셀 원본 파일에 성공적으로 저장되었습니다! 화면을 새로고침 합니다.")
                        st.rerun()
                    else:
                        st.warning("저장할 변경사항이 없습니다.")
                except Exception as save_err:
                    st.error(f"엑셀 저장 중 오류가 발생했습니다: {save_err}")

        else:
            st.warning(f"선택하신 조건에 해당하는 ECN 내역이 없습니다.")
            
    except Exception as e:
        if "404" in str(e):
            st.error(f"⚠️ 깃허브에 **`{target_file}`** 파일이 없습니다. 엑셀 양식을 만들어 업로드해주세요.")
        else:
            st.error(f"⚠️ 파일을 읽는 중 오류가 발생했습니다: {e}")

# ==========================================
# 7. 메인 실행 
# ==========================================
def main():
    try:
        g = Github(st.secrets["GITHUB_TOKEN"]); repo = g.get_repo(st.secrets["REPO_NAME"])
        db_log = DataManager(repo, st.secrets["FILE_PATH"], ["날짜", "장비", "작성자", "업무내용", "비고", "첨부"])
        db_flow = DataManager(repo, "cs_flow_data.csv", ["프로젝트명", "대항목", "작업내용", "상태", "비고", "첨부", "업데이트일"])
    except Exception as e:
        st.error(f"⚠️ 깃허브 연결 설정 오류: {e}")
        return

    if 'logged_in' not in st.session_state: st.session_state.update({'logged_in': False, 'user_name': ""})

    if not st.session_state['logged_in']:
        with st.form("login"):
            name = st.text_input("성함"); 
            if st.form_submit_button("입장") and name: 
                st.session_state.update({'logged_in': True, 'user_name': name}); st.rerun()
    else:
        top_col1, top_col2, top_col3 = st.columns([6, 3, 1])
        with top_col2:
            menu_selection = st.selectbox("📂 대시보드 메뉴 이동", ["📝 업무일지", "✅ CS 작업체크시트", "📊 장비가동데이터", "🛠️ ECN & STN"], label_visibility="collapsed")
        with top_col3:
            if st.button("🚪 로그아웃", use_container_width=True): 
                st.session_state['logged_in'] = False
                st.rerun()

        st.sidebar.markdown(f"👤 **{st.session_state['user_name']}** 님 환영합니다.")

        if menu_selection == "📝 업무일지": render_work_log_page(db_log)
        elif menu_selection == "✅ CS 작업체크시트": render_cs_flow_page(db_flow)
        elif menu_selection == "📊 장비가동데이터": render_equipment_data_page()
        elif menu_selection == "🛠️ ECN & STN": render_ecn_stn_page(repo)

if __name__ == "__main__": main()
